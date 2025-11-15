import subprocess
import json
import pathlib
import openpyxl
import datetime
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# --- Configuration ---
# Adjust these paths if your benchmark names are different!
PERFT_JSON_PATH = "C:/Users/Moritz/RustroverProjects/ChessEngine/target/criterion/standard_perft5/new/estimates.json"
EVAL_JSON_PATH = "C:/Users/Moritz/RustroverProjects/ChessEngine/target/criterion/standard_board_evaluation/new/estimates.json"
EXCEL_FILE = "C:/Users/Moritz/RustroverProjects/ChessEngine/progress_tracking/progress.xlsx"
HEADERS = ["TIMESTAMP", "COMMIT", "MESSAGE", "PERFT (ms)", "EVAL (ps)", "SUITE (%)"]

COLUMN_WIDTHS = {
    'A': 20,  # Timestamp
    'B': 12,  # Commit
    'C': 50,  # Message
    'D': 14,  # Perft
    'E': 14,  # Eval
    'F': 14   # Suite
}

# NEW: Define fonts
DEFAULT_FONT = Font(name='Consolas', size=11)
HEADER_FONT = Font(name='Consolas', size=11, bold=True)
# ---------------------

def run_command(command):
    """Executes a shell command and returns its output."""
    print(f"Running: {' '.join(command)}")
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True, encoding='utf-8')
        return result
    except subprocess.CalledProcessError as e:
        print(f"Error running command: {e}")
        print("STDOUT:", e.stdout)
        print("STDERR:", e.stderr)
        exit(1)

def get_criterion_result(json_path):
    """Reads the result from a Criterion JSON file."""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # Returns the 'point_estimate' of the mean in nanoseconds
        return data['mean']['point_estimate']
    except FileNotFoundError:
        print(f"Error: JSON file not found: {json_path}")
        print("Make sure 'cargo bench' was successful and the paths are correct.")
        exit(1)
    except (KeyError, TypeError):
        print(f"Error: Unexpected format in {json_path}")
        exit(1)

def get_git_info():
    """Checks if the git working directory is dirty. Returns (hash, message)"""
    status_result = run_command(["git", "status", "--porcelain"])

    if status_result.stdout.strip():
        print("Uncommitted changes detected. Using 'local' as commit ID.")
        return ("local", "Uncommitted changes")
    else:
        hash_result = run_command(["git", "rev-parse", "--short", "HEAD"])
        msg_result = run_command(["git", "log", "-1", "--pretty=%s"])
        return (hash_result.stdout.strip(), msg_result.stdout.strip())

def apply_styles_and_formats(ws, row_index, is_header=False):
    """Applies fonts and number formats to a specific row."""
    font = HEADER_FONT if is_header else DEFAULT_FONT

    # Get column indices
    try:
        perft_col_idx = HEADERS.index('PERFT (ms)') + 1
        eval_col_idx = HEADERS.index('EVAL (ps)') + 1
        suite_col_idx = HEADERS.index('SUITE (%)') + 1
    except ValueError:
        print("Error: Could not find all headers. Check HEADERS config.")
        return

    for cell in ws[row_index]:
        cell.font = font

        # Apply number formats only to data rows
        if not is_header:
            if cell.column == perft_col_idx or cell.column == eval_col_idx or cell.column == suite_col_idx:
                cell.number_format = '0.00'

def main():
    # 1. Run benchmarks and suite
    print("Starting benchmarks... (This may take a few minutes)")
    run_command(["cargo", "bench", "--bench", "perft"])
    run_command(["cargo", "bench", "--bench", "eval"])

    print("Starting suite test...")
    suite_result = run_command(["cargo", "run", "--bin", "suite", "--release"])

    try:
        # The suite_score is still a raw float, e.g., 95.5
        suite_score = float(suite_result.stdout.strip())
    except ValueError:
        print(f"Error: Could not convert suite output to a number.")
        print(f"Received: '{suite_result.stdout}'")
        exit(1)

    print("Collecting results...")

    # 2. Get Git info and Timestamp
    (commit_hash, commit_message) = get_git_info()
    timestamp = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

    # 3. Read benchmark results
    # Convert from nanoseconds to milliseconds
    perft_ms = get_criterion_result(PERFT_JSON_PATH) / 1_000_000.0
    # Convert from nanoseconds to picoseconds
    eval_ps = get_criterion_result(EVAL_JSON_PATH) * 1000.0

    # 4. Write data to the Excel file
    file_path = pathlib.Path(EXCEL_FILE)

    if file_path.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        # Check if cell A1 has the correct header. If not, the file is empty/corrupt
        if ws.cell(row=1, column=1).value != HEADERS[0]:
            print("File was empty or corrupt. Re-creating headers.")
            ws.append(HEADERS)
            apply_styles_and_formats(ws, 1, is_header=True)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Progress"
        ws.append(HEADERS)
        apply_styles_and_formats(ws, 1, is_header=True) # Apply header style
        print(f"New file '{EXCEL_FILE}' created.")

    # --- Set Column Widths ---
    # !! This was the fix: Removed the "if" check and adjusted units.
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Overwrite Logic ---
    if commit_hash == "local" and ws.max_row > 1:
        try:
            commit_col_index = HEADERS.index("COMMIT") + 1
        except ValueError:
            print("Error: 'COMMIT' column not found in headers.")
            exit(1)

        last_row_commit_val = ws.cell(row=ws.max_row, column=commit_col_index).value

        if last_row_commit_val == "local":
            ws.delete_rows(ws.max_row)
            print("Overwriting previous 'local' entry.")

    # Append the new row of data (using ms values)
    new_row = [timestamp, commit_hash, commit_message, perft_ms, eval_ps, suite_score]
    ws.append(new_row)

    # Apply default font and number formats to the newly added row
    apply_styles_and_formats(ws, ws.max_row, is_header=False)


    # --- Add/Update Conditional Formatting ---
    perf_rule = ColorScaleRule(
        start_type='min', start_color='63BE7B', # Green (Low = Fast = Good)
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B' # Red (High = Slow = Bad)
    )
    suite_rule = ColorScaleRule(
        start_type='min', start_color='F8696B', # Red (Low = Bad)
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='63BE7B' # Green (High = Good)
    )

    try:
        perft_col_letter = get_column_letter(HEADERS.index('PERFT (ms)') + 1)
        # Note: This had a typo in your original file 'EVAL (fs)', I assume you meant 'EVAL (ps)'
        eval_col_letter = get_column_letter(HEADERS.index('EVAL (ps)') + 1)
        suite_col_letter = get_column_letter(HEADERS.index('SUITE (%)') + 1)

        max_excel_row = 1048576 # Standard for .xlsx
        ws.conditional_formatting.add(f'{perft_col_letter}2:{perft_col_letter}{max_excel_row}', perf_rule)
        ws.conditional_formatting.add(f'{eval_col_letter}2:{eval_col_letter}{max_excel_row}', perf_rule)
        ws.conditional_formatting.add(f'{suite_col_letter}2:{suite_col_letter}{max_excel_row}', suite_rule)

    except ValueError:
        print("Warning: Could not find performance columns in headers. Skipping color formatting.")
        # Print which headers are problematic
        for col in ['PERFT (ms)', 'EVAL (ps)', 'SUITE (%)']:
            if col not in HEADERS:
                print(f"Header '{col}' is missing or misspelled in HEADERS list.")


    # 5. Save the file
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        print(f"Error: Could not save '{EXCEL_FILE}'.")
        print("Please make sure the file is not open in Excel.")
        exit(1)

    print("-" * 30)
    print(f"Success! Results saved to '{EXCEL_FILE}'.")
    print(f"  TIMESTAMP: {timestamp}")
    print(f"  COMMIT:    {commit_hash}")
    print(f"  MESSAGE:   {commit_message}")
    print(f"  PERFT:     {perft_ms:.2f} ms")
    print(f"  EVAL:      {eval_ps:.2f} ps")
    print(f"  SUITE:     {suite_score:.2f} %")

if __name__ == "__main__":
    main()